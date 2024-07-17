import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook and select the sheet
workbook_path = 'your_excel_file.xlsx'
column_letter = 'A'  # Change this to the letter of the column you are interested in

wb = openpyxl.load_workbook(workbook_path)
sheet = wb.active

# Define the RGB color code for yellow
yellow_rgb = 'FFFF00'

# Find the column letter based on the column name in the header row
column_letter = None
for cell in sheet[1]:  # Assuming the first row contains the column headers
    if cell.value == column_name:
        column_letter = cell.column_letter
        break

if not column_letter:
    raise ValueError(f"Column '{column_name}' not found in the header row.")

highlighted_rows = []

# Iterate through the rows in the specified column
for row in range(1, sheet.max_row + 1):
    cell = sheet[f"{column_letter}{row}"]
    fill = cell.fill
    if fill and fill.start_color:
        fg_color = fill.start_color
        if fg_color.type == 'rgb' and fg_color.rgb == yellow_rgb:
            highlighted_rows.append(row)
        elif fg_color.type == 'indexed' and fg_color.index == 5:  # Yellow index
            highlighted_rows.append(row)

# Print or process the highlighted rows
print("Rows with yellow highlights in column", column_letter, ":", highlighted_rows)


# Read the .dat file and extract FILE_NO values
file_no_values = []
with open(dat_file_path, 'r') as file:
    for line in file:
        parts = line.strip().split()
        if len(parts) >= 3 and parts[2].startswith("FILE_NO"):
            file_no_values.append(parts[2])

# Find the matching values
matching_values = [value for value in highlighted_values if value in file_no_values]

# Print the matching values
print("Matching values:", matching_values)





# Define the file paths
may_file_path = '/mnt/data/may_resources.txt'
june_file_path = '/mnt/data/june_resources.txt'

# Read the contents of the May resources file
with open(may_file_path, 'r') as may_file:
    may_resources = set(may_file.read().splitlines())

# Read the contents of the June resources file
with open(june_file_path, 'r') as june_file:
    june_resources = set(june_file.read().splitlines())

# Find resources that are in May but not in June
missing_in_june = may_resources - june_resources

# Print the missing resources
print("Resources present in May but missing in June:")
for resource in missing_in_june:
    print(resource)
